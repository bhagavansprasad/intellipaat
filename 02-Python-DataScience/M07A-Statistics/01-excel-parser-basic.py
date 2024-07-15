from openpyxl import load_workbook
from openpyxl.styles import Font, Fill, colors, PatternFill


def print_all_worksheet_names(wbname):
    wb = load_workbook(wbname)

    #print worksheet names
    print(wb.sheetnames)

    for wsheet in wb.sheetnames:
        print(wsheet)
    wb.close()
    print ("")

def print_cell_values(wbname):
    wb = load_workbook(wbname)

    # opening worksheet in workbook
    wsheet = wb['sales']

    # max rows and cols
    rcount =  wsheet.max_row
    ccount =  wsheet.max_column
    print("max rows filled :", rcount)
    print("max cols filled :", ccount)
    print ("")
    
    #print cell value
    print("A2           :", wsheet['A2'].value) 
    print("C3           :", wsheet['C3'].value) 
    print ("")

    #print in row and col format
    print("wsheet[2][0] :", wsheet[2][0].value) 
    print("wsheet[3][2] :", wsheet[3][2].value) 
    print ("")

    wb.close()
    return

def print_rows_columns1(wbname):
    #opening workbook
    wb = load_workbook(wbname)

    #opening worksheet in workbook
    wsheet = wb['sales']

    #max rows and cols
    rcount =  wsheet.max_row
    ccount =  wsheet.max_column

    #print rows and columns
    for i in range(1, rcount+1):
        for j in range (0, ccount):
            print((wsheet[i][j].value), end=' ')
        print("")

    wb.close()
    print ("")

def print_rows_columns2(wbname):
    #opening workbook
    wb = load_workbook(wbname)

    #opening worksheet in workbook
    wsheet = wb['sales']

    #max rows and cols
    rcount =  wsheet.max_row
    ccount =  wsheet.max_column

    #print rows and columns
    for row in range(1, rcount+1):
        if not any(cell.value for cell in wsheet[row]):
            print("The %d row is empty" % row)
            break

        for col in range (0, ccount):
            if (wsheet[row][col].value != None):
                print((str(wsheet[row][col].value)), end=' ')

        print ("")
    wb.close()


def highlight_cell_by_value(wbname, wsheet_name, cell_value):
    #opening workbook
    wb = load_workbook(wbname)

    #opening worksheet in workbook
    wsheet = wb[wsheet_name]

    #max rows and cols
    rcount =  wsheet.max_row
    ccount =  wsheet.max_column

    #Highlight cell, if the cell value is Ashish
    #and change cell font size to 15, color GREEN, bold

    #print (type(wsheet[1]))
    #print (wsheet[1])

    for row in range(1, rcount+1):
        if not any(cell.value for cell in wsheet[row]):
            print("The %d row is empty" % row)
            break

        for col in range (0, ccount):
            if (wsheet[row][col].value != None and wsheet[row][col].value == cell_value):
                #print("font :", wsheet[row][col].font)
                wsheet[row][col].font = Font(size=15, bold=True, color=colors.BLACK)
                print("")

    wb.save(wbname)
    wb.close()

#Highlight row background if the cell value is as cell_value
#Change row background to yellow
def change_background_cell_by_value(wbname, wsheet_name, cell_value):
    #opening workbook
    wb = load_workbook(wbname)

    #opening worksheet in workbook
    wsheet = wb[wsheet_name]

    #max rows and cols
    rcount =  wsheet.max_row
    ccount =  wsheet.max_column

    for row in range(1, rcount):
        if not any(cell.value for cell in wsheet[row]):
            print("The %d row is empty" % row)
            break

        for col in range (0, ccount):
            if (wsheet[row][col].value != None and wsheet[row][col].value == cell_value ):
                print(wsheet[row][col].fill)
                print(str(wsheet[row][col].value))

                #for cell in wsheet[row:row]:
                for cell in wsheet[row]:
                    cell.fill = PatternFill(fill_type="solid", start_color=colors.BLUE)

    wb.save(wbname)
    wb.close()

def main():
    filename = "revenue.xlsx"
    wsheet_name = "sales"
    tstr = "Ashish"
    # print_all_worksheet_names(filename)
    # print_cell_values(filename)
    # print_rows_columns1(filename)
    # print_rows_columns2(filename)
    # print(colors)
    # print(dir(colors))
    # print(help(colors))
    # highlight_cell_by_value(filename, wsheet_name, tstr)

    tstr = "Kavitha"
    change_background_cell_by_value(filename, wsheet_name, tstr)
    return

if (__name__ == '__main__'):
    main()

